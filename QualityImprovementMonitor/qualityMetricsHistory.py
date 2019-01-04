# Encapsulates warning setting history Excel report
# The report is updated with data from the warning settings report

import os
from openpyxl import *
from datetime import *


class QualityMetricsHistory:

    def __init__(self, unit_collection):
        self.unit_collection = unit_collection
        self.filename = r'QualityMetricsHistory.xlsx'
        self.worksheet_names = ["Wrong warning level", "Treat warning not as error", "Suppressed warnings", "Actual warnings", "Coverity level 1", "Coverity level 2", "Security level 1", "Security level 2"]

        self.date_format = "%d-%b-%Y"

        if not os.path.isfile(self.filename):
            self._create_workbook()

    def update(self):
        print("begin update history")

        today = date.today()
        time_stamp = today.strftime(self.date_format)

        for unit in self.unit_collection.units:
            unit.update_path_last_successful_build()

        wb = load_workbook(self.filename)
        self._fill_worksheet_wrong_warning_level(wb.worksheets[0], time_stamp)
        self._fill_worksheet_treat_warnings_not_as_errors(wb.worksheets[1], time_stamp)
        self._fill_worksheet_suppressed_warnings(wb.worksheets[2], time_stamp)
        self._fill_worksheet_actual_warnings(wb.worksheets[3], time_stamp)
        self._fill_worksheet_coverity_warnings(wb.worksheets[4], time_stamp, 1)
        self._fill_worksheet_coverity_warnings(wb.worksheets[5], time_stamp, 2)
        self._fill_worksheet_security_warnings(wb.worksheets[6], time_stamp, 1)
        self._fill_worksheet_security_warnings(wb.worksheets[7], time_stamp, 2)

        wb.save(self.filename)

        print("end update history")

    # wrong warning level
    def get_values_wrong_warning_level(self):
        wb = load_workbook(self.filename)
        return self._get_values(wb.worksheets[0])

    def get_deltas_wrong_warning_level(self):
        wb = load_workbook(self.filename)
        return self._get_deltas(wb.worksheets[0])

    def get_history_wrong_warning_level(self):
        wb = load_workbook(self.filename)
        return self._get_history(wb.worksheets[0])

    # treat warning not as errors
    def get_values_treat_warnings_not_as_errors(self):
        wb = load_workbook(self.filename)
        return self._get_values(wb.worksheets[1])

    def get_deltas_treat_warnings_not_as_errors(self):
        wb = load_workbook(self.filename)
        return self._get_deltas(wb.worksheets[1])

    def get_history_treat_warnings_not_as_errors(self):
        wb = load_workbook(self.filename)
        return self._get_history(wb.worksheets[1])

    # suppressed warnings
    def get_values_suppressed_warnings(self):
        wb = load_workbook(self.filename)
        return self._get_values(wb.worksheets[2])

    def get_deltas_suppressed_warnings(self):
        wb = load_workbook(self.filename)
        return self._get_deltas(wb.worksheets[2])

    def get_history_suppressed_warnings(self):
        wb = load_workbook(self.filename)
        return self._get_history(wb.worksheets[2])

    # actual warnings
    def get_values_actual_warnings(self):
        wb = load_workbook(self.filename)
        return self._get_values(wb.worksheets[3])

    def get_deltas_actual_warnings(self):
        wb = load_workbook(self.filename)
        return self._get_deltas(wb.worksheets[3])

    def get_history_actual_warnings(self):
        wb = load_workbook(self.filename)
        return self._get_history(wb.worksheets[3])

    # warnings metric
    def get_history_warning_suppression_indicator(self):
        totals = {}

        history_wrong_warning_level = self.get_history_wrong_warning_level()
        history_treat_warnings_not_as_errors = self.get_history_treat_warnings_not_as_errors()
        history_suppressed_warnings = self.get_history_suppressed_warnings()
        history_actual_warnings = self.get_history_actual_warnings()

        weight_treat_warnings_not_as_errors = 0
        weight_wrong_warning_levels = 0
        weight_suppressed_warnings = 1
        weight_actual_warnings = 1

        for key, value in history_wrong_warning_level.items():
            total = (history_treat_warnings_not_as_errors[key] * weight_treat_warnings_not_as_errors) + \
                    (history_wrong_warning_level[key]) * weight_wrong_warning_levels +\
                    (history_suppressed_warnings[key] * weight_suppressed_warnings) + \
                    (history_actual_warnings[key] * weight_actual_warnings)

            totals[key] = total

        return totals

    # coverity
    def get_values_coverity_level_1(self):
        wb = load_workbook(self.filename)
        return self._get_values(wb.worksheets[4])

    def get_deltas_coverity_level_1(self):
        wb = load_workbook(self.filename)
        return self._get_deltas(wb.worksheets[4])

    def get_history_coverity_level_1(self):
        wb = load_workbook(self.filename)
        return self._get_history(wb.worksheets[4])

    def get_values_coverity_level_2(self):
        wb = load_workbook(self.filename)
        return self._get_values(wb.worksheets[5])

    def get_deltas_coverity_level_2(self):
        wb = load_workbook(self.filename)
        return self._get_deltas(wb.worksheets[5])

    def get_history_coverity_level_2(self):
        wb = load_workbook(self.filename)
        return self._get_history(wb.worksheets[5])

    # security
    def get_values_security_level_1(self):
        wb = load_workbook(self.filename)
        return self._get_values(wb.worksheets[6])

    def get_deltas_security_level_1(self):
        wb = load_workbook(self.filename)
        return self._get_deltas(wb.worksheets[6])

    def get_history_security_level_1(self):
        wb = load_workbook(self.filename)
        return self._get_history(wb.worksheets[6])

    def get_values_security_level_2(self):
        wb = load_workbook(self.filename)
        return self._get_values(wb.worksheets[7])

    def get_deltas_security_level_2(self):
        wb = load_workbook(self.filename)
        return self._get_deltas(wb.worksheets[7])

    def get_history_security_level_2(self):
        wb = load_workbook(self.filename)
        return self._get_history(wb.worksheets[7])

    def _create_workbook(self):
        wb = Workbook()

        for index in range(8):
            if index != 0:
                wb.create_sheet()

            self._fill_worksheet_headers(wb.worksheets[index], self.worksheet_names[index])

        wb.save(self.filename)

    def _fill_worksheet_headers(self, worksheet, title):
        worksheet.title = title

        current_column = 1

        worksheet.cell(1, current_column).value = "Date"
        current_column += 1
        for unit in self.unit_collection.units:
            worksheet.cell(1, current_column).value = unit.unit_name
            current_column += 1
        worksheet.cell(1, current_column).value = "Total"

    def _fill_worksheet_wrong_warning_level(self, worksheet, time_stamp):
        count_for_all_units = []
        for unit in self.unit_collection.units:
            count_for_unit = unit.get_latest_build_wrong_warning_level_count()
            count_for_all_units.append(count_for_unit)
        self._fill_worksheet(worksheet, time_stamp, count_for_all_units)

    def _fill_worksheet_treat_warnings_not_as_errors(self, worksheet, time_stamp):
        count_for_all_units = []
        for unit in self.unit_collection.units:
            count_for_unit = unit.get_latest_build_treat_warnings_not_as_errors_count()
            count_for_all_units.append(count_for_unit)
        self._fill_worksheet(worksheet, time_stamp, count_for_all_units)

    def _fill_worksheet_suppressed_warnings(self, worksheet, time_stamp):
        count_for_all_units = []
        for unit in self.unit_collection.units:
            count_for_unit = unit.get_latest_build_suppressed_warning_count()
            count_for_all_units.append(count_for_unit)
        self._fill_worksheet(worksheet, time_stamp, count_for_all_units)

    def _fill_worksheet_actual_warnings(self, worksheet, time_stamp):
        count_for_all_units = []
        for unit in self.unit_collection.units:
            count_for_unit = unit.get_latest_build_actual_warning_count()
            count_for_all_units.append(count_for_unit)
        self._fill_worksheet(worksheet, time_stamp, count_for_all_units)

    def _fill_worksheet_coverity_warnings(self, worksheet, time_stamp, level):
        count_for_all_units = []
        for unit in self.unit_collection.units:
            count_for_unit = unit.get_latest_build_coverity_error_count(level)
            count_for_all_units.append(count_for_unit)

        self._fill_worksheet(worksheet, time_stamp, count_for_all_units)

    def _fill_worksheet_security_warnings(self, worksheet, time_stamp, level):
        count_for_all_units = []
        for unit in self.unit_collection.units:
            count_for_unit = unit.get_latest_build_security_error_count(level)
            count_for_all_units.append(count_for_unit)

        self._fill_worksheet(worksheet, time_stamp, count_for_all_units)

    def _fill_worksheet(self, worksheet, time_stamp, count_for_all_units):
        current_column = 1
        current_row = worksheet.max_row

        last_time_stamp = worksheet.cell(current_row, 1).value
        if not last_time_stamp == time_stamp:
            current_row += 1

        total = 0
        worksheet.cell(current_row, current_column).value = time_stamp
        current_column += 1
        for count in count_for_all_units:
            total += count
            worksheet.cell(current_row, current_column).value = count
            current_column += 1
        worksheet.cell(current_row, current_column).value = total

    def _get_values(self, worksheet):
        values = []
        current_row = worksheet.max_row

        for index in range(len(self.unit_collection.units)+1):
            current_column = index + 2
            actual_value = worksheet.cell(current_row, current_column).value
            values.append(actual_value)
            current_column += 1
        return values

    def _get_deltas(self, worksheet):
        deltas = []
        current_row = worksheet.max_row

        for index in range(len(self.unit_collection.units)+1):
            if current_row > 2:
                current_column = index + 2
                actual_value = worksheet.cell(current_row, current_column).value

                delta = 0
                days_look_back = 1

                row = current_row
                while delta == 0 and row > 2 and days_look_back > 0:
                    row -= 1
                    days_look_back -= 1
                    previous_value = worksheet.cell(row, current_column).value
                    delta = actual_value - previous_value

                deltas.append(delta)
            else:
                deltas.append(0)
        return deltas

    def _get_history(self, worksheet):
        totals = {}
        date_column = 1
        totals_column = len(self.unit_collection.units) + 2

        for row in range(2, worksheet.max_row):
            cell_value = worksheet.cell(row, date_column).value

            # convert to datetime if date stored as string in excel sheet
            if isinstance(cell_value, str):
                date = datetime.strptime(cell_value, self.date_format)
            else:
                date = cell_value

            total = worksheet.cell(row, totals_column).value
            totals[date] = total
        return totals